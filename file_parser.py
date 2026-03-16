import openpyxl
from datetime import datetime

def parse_missing_cdf(file_path, report_date):
    """
    Count new missing CDFs from the MissingCDF tab
    report_date format: YYYYMMDD e.g. 20260303
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Check if MissingCDF tab exists
        if 'MissingCDF' not in wb.sheetnames:
            print(f"MissingCDF tab not found in {file_path}")
            print(f"Available tabs: {wb.sheetnames}")
            return 0
        
        ws = wb['MissingCDF']
        
        # Convert report_date to match the format in the file
        # File format appears to be YYYY-MM-DD e.g. 2026-03-03
        report_date_formatted = f"{report_date[:4]}-{report_date[4:6]}-{report_date[6:8]}"
        
        # Find the header row and IMAGE_FILE_PROCESSING_DATE column
        header_row = None
        date_col = None
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'IMAGE_FILE_PROCESSING_DATE':
                    header_row = cell.row
                    date_col = cell.column
                    break
            if header_row:
                break
        
        if not header_row or not date_col:
            print("Could not find IMAGE_FILE_PROCESSING_DATE column")
            return 0
        
        # Count rows where date matches report date
        new_cdf_count = 0
        for row in ws.iter_rows(min_row=header_row + 1):
            cell_value = row[date_col - 1].value
            
            if cell_value is None:
                continue
            
            # Handle both string and datetime formats
            if isinstance(cell_value, datetime):
                cell_date = cell_value.strftime("%Y-%m-%d")
            else:
                cell_date = str(cell_value).strip()[:10]
            
            if cell_date == report_date_formatted:
                new_cdf_count += 1
        
        print(f"Found {new_cdf_count} new CDFs for date {report_date_formatted}")
        return new_cdf_count
    
    except Exception as e:
        print(f"Error parsing MissingCDF file: {e}")
        return 0

def parse_missing_images(file_path):
    """
    Count new failed images from the Detail_w_Errors tab
    New = AGE_IN_WEEKS = 0 and IMG_STATUS = FAILED
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Check if Detail_w_Errors tab exists
        if 'Detail_w_Errors' not in wb.sheetnames:
            print(f"Detail_w_Errors tab not found in {file_path}")
            print(f"Available tabs: {wb.sheetnames}")
            return 0
        
        ws = wb['Detail_w_Errors']
        
        # Find header row and relevant columns
        header_row = None
        status_col = None
        age_weeks_col = None
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'IMG_STATUS':
                    header_row = cell.row
                    status_col = cell.column
                if cell.value == 'AGE_IN_WEEKS':
                    age_weeks_col = cell.column
            if header_row:
                break
        
        if not header_row or not status_col or not age_weeks_col:
            print("Could not find required columns in Detail_w_Errors tab")
            return 0
        
        # Count rows where IMG_STATUS = FAILED and AGE_IN_WEEKS = 0
        new_failed_count = 0
        for row in ws.iter_rows(min_row=header_row + 1):
            status = row[status_col - 1].value
            age_weeks = row[age_weeks_col - 1].value
            
            if status is None:
                continue
            
            if str(status).strip().upper() == 'FAILED' and str(age_weeks).strip() == '0':
                new_failed_count += 1
        
        print(f"Found {new_failed_count} new failed images")
        return new_failed_count
    
    except Exception as e:
        print(f"Error parsing MissingImages file: {e}")
        return 0

if __name__ == "__main__":
    # Test - download files and parse them
    from s3_handler import get_file_for_error
    
    test_date = "20260303"
    
    print("Testing MissingCDF parser...")
    cdf_file = get_file_for_error("MISSING_CDF", test_date)
    if cdf_file:
        count = parse_missing_cdf(cdf_file, test_date)
        print(f"New CDFs: {count}")
    
    print("\nTesting MissingImages parser...")
    images_file = get_file_for_error("FAILED_IMAGES", test_date)
    if images_file:
        count = parse_missing_images(images_file)
        print(f"New Failed Images: {count}")